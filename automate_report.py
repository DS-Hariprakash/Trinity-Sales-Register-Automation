# ============================================================
# automate_report.py
# Trinity Sales Register — Ramco ERP Sales Register Mattresses
# Automation (download only).
#
# Same navigation/export flow as the Sales Register project's
# automate_report_v1.py, but intended for unattended 7 AM runs:
#   • headless by default (configurable)
#   • no "visual confirmation" pauses
#   • the post-download step is EMAIL (see email_sender.py),
#     NOT the WhatsApp breakup message.
# The downloaded .xlsx is shared exactly as exported — no edits.
# ============================================================

import os
import re
import sys
import time
import zipfile
from datetime import datetime, date
from calendar import monthrange
from pathlib import Path

from dotenv import load_dotenv
from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeout
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


sys.stdout.reconfigure(encoding="utf-8")  # type: ignore[union-attr]
sys.stderr.reconfigure(encoding="utf-8")  # type: ignore[union-attr]

# ── Config ───────────────────────────────────────────────────
load_dotenv(dotenv_path=Path(__file__).parent / "config.env")


def _require_env(name: str) -> str:
    value = os.getenv(name)
    if not value:
        raise RuntimeError(f"Missing required setting '{name}' in config.env")
    return value


ERP_URL        = _require_env("ERP_URL")
USERNAME       = _require_env("ERP_USERNAME")
PASSWORD       = _require_env("ERP_PASSWORD")
DOWNLOAD_DIR   = Path(_require_env("DOWNLOAD_DIR"))
DURATION_FILE  = Path(os.getenv("DURATION_LOG_FILE", "")) if os.getenv("DURATION_LOG_FILE") else None

# Headless for unattended scheduled runs; flip to False to watch/debug.
_HEADLESS_RAW  = os.getenv("HEADLESS", "True").strip().lower()
HEADLESS       = _HEADLESS_RAW in ("1", "true", "yes", "on")
# Brief settle time in place of the old 30s visual-confirmation pauses.
SETTLE_SECS    = float(os.getenv("SETTLE_SECS", "2"))

REPORT_NAME    = "Sales Register Mattresses"
REPORT_SNUM    = 26   # S.No in the duration Excel
TODAY          = date.today()
SHEET_NAME     = TODAY.strftime("%d-%m-%Y")   # e.g. 24-06-2026

# Current month date range
MONTH_FROM     = date(TODAY.year, TODAY.month, 1)
MONTH_TO       = date(TODAY.year, TODAY.month, monthrange(TODAY.year, TODAY.month)[1])
DATE_FMT       = "%d-%m-%Y"   # RamcoHub date fields use DD-MM-YYYY

# Set to True to narrow to a 2-day window for fast iteration while
# debugging the Export flow. Must be False for real scheduled runs.
TEST_MODE = False
if TEST_MODE:
    MONTH_FROM = date(TODAY.year, TODAY.month, 1)
    MONTH_TO = date(TODAY.year, TODAY.month, 2)


# ── Helpers ──────────────────────────────────────────────────

def fix_xlsx_relationship_paths(path: Path):
    """
    Repairs two defects in this ERP's export generator, both invisible to
    openpyxl but enough for Excel itself to refuse the file ("file format
    or file extension is not valid"):

    1. One worksheet relationship in xl/_rels/workbook.xml.rels uses an
       absolute target path (e.g. "/xl/worksheets/sheet4.xml") while its
       siblings use relative ones ("worksheets/sheet3.xml").
    2. [Content_Types].xml declares the workbook part as macro-enabled
       (the .xlsm content type) even though there's no vbaProject.bin and
       the file is saved as .xlsx.
    """
    rels_path = "xl/_rels/workbook.xml.rels"
    content_types_path = "[Content_Types].xml"
    with zipfile.ZipFile(path, "r") as zin:
        items = zin.infolist()
        data = {item.filename: zin.read(item.filename) for item in items}

    changed = False

    rels = data[rels_path].decode("utf-8")
    fixed_rels = re.sub(r'Target="/xl/(worksheets/[^"]+)"', r'Target="\1"', rels)
    if fixed_rels != rels:
        data[rels_path] = fixed_rels.encode("utf-8")
        changed = True

    content_types = data[content_types_path].decode("utf-8")
    fixed_content_types = content_types.replace(
        "application/vnd.ms-excel.sheet.macroEnabled.main+xml",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml",
    )
    if fixed_content_types != content_types:
        data[content_types_path] = fixed_content_types.encode("utf-8")
        changed = True

    if not changed:
        return

    with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED) as zout:
        for item in items:
            zout.writestr(item, data[item.filename])


def format_duration(seconds: float) -> str:
    """Convert seconds to human-readable string: 3 Min 42 Sec"""
    m = int(seconds // 60)
    s = int(seconds % 60)
    if m > 0:
        return f"{m} Min {s} Sec"
    return f"{s} Sec"


def wait_ready(page, timeout: int = 60000):
    """
    Waits for network to settle and for the RamcoHub 'Please Wait' load
    mask to clear. The ERP app re-shows this mask on every screen
    transition, and clicking while it's up gets intercepted.
    """
    page.wait_for_load_state("networkidle", timeout=timeout)
    try:
        page.wait_for_selector("text=Please Wait", state="hidden", timeout=timeout)
    except PlaywrightTimeout:
        pass
    page.wait_for_timeout(1000)


_FIND_VISIBLE_TEXT_JS = """(args) => {
    const { selector, label, exclude, attr } = args;
    const all = document.querySelectorAll(selector);
    let fallback = null;
    for (const el of all) {
        const rect = el.getBoundingClientRect();
        const style = getComputedStyle(el);
        const visible = rect.width > 0 && rect.height > 0
            && style.visibility !== 'hidden' && style.display !== 'none';
        if (!visible || el.textContent.trim() !== label) continue;
        const cls = el.className || '';
        if (exclude.some(bad => cls.includes(bad))) {
            if (!fallback) fallback = el;
            continue;
        }
        el.setAttribute(attr, 'true');
        return true;
    }
    if (fallback) {
        fallback.setAttribute(attr, 'true');
        return true;
    }
    return false;
}"""


def wait_for_visible_text(target, selector: str, label: str, timeout: int = 150000,
                           interval: int = 500, exclude=None):
    """
    Finds a visible element with exact text, doing the whole scan in a
    single JS round-trip per poll, and returns it as a real Playwright
    locator for the caller to interact with (e.g. .click()) — using
    genuine mouse events matters here: ExtJS form fields commit their
    typed value to the underlying model on blur, and a synthetic JS
    click() on the next button never fires that blur on whatever field
    was focused before it.

    `target` is anything with .evaluate()/.locator() — a Page or a Frame
    (RamcoHub renders report dialogs inside an iframe, so callers must
    pass the right one). RamcoHub's SPA keeps hundreds of hidden/cached
    widgets in the DOM, so checking each Playwright locator match one at
    a time (one round-trip per element, every poll) never completes a
    full pass within a sane timeout. CSS text-selector engines (text=,
    :text-is(), :has-text()) also proved unreliable against this app's
    markup. Doing the search natively via evaluate() is fast and robust.
    """
    deadline = time.time() + timeout / 1000
    attr = "data-claude-target2"
    args = {"selector": selector, "label": label, "exclude": exclude or [], "attr": attr}
    while time.time() < deadline:
        if target.evaluate(_FIND_VISIBLE_TEXT_JS, args):
            return target.locator(f"[{attr}='true']")
        time.sleep(interval / 1000)
    raise PlaywrightTimeout(f"No visible element with text '{label}' found for '{selector}' within {timeout}ms")


def find_visible_input_by_value(target, pattern: str, placeholder: str | None = None):
    """
    Returns a locator for the first visible <input> whose current value
    matches a regex (an already-filled date) or whose placeholder
    attribute matches exactly (the empty "DD-MM-YYYY" state). `target` is
    the Page/Frame the field lives in — see wait_for_visible_text.
    """
    js = """(args) => {
        const { pattern, placeholder, attr } = args;
        const re = new RegExp(pattern);
        for (const el of document.querySelectorAll('input')) {
            const rect = el.getBoundingClientRect();
            const style = getComputedStyle(el);
            const visible = rect.width > 0 && rect.height > 0
                && style.visibility !== 'hidden' && style.display !== 'none';
            if (!visible) continue;
            if (re.test(el.value) || (placeholder && el.placeholder === placeholder)) {
                el.setAttribute(attr, 'true');
                return true;
            }
        }
        return false;
    }"""
    attr = "data-claude-target"
    found = target.evaluate(js, {"pattern": pattern, "placeholder": placeholder, "attr": attr})
    return target.locator(f"input[{attr}='true']") if found else None


def _parse_duration_seconds(duration_str):
    """Parses 'X Min Y Sec' / 'Y Sec' (and tolerantly 'X Mins') into seconds."""
    if not duration_str:
        return None
    m = re.search(r'(?:(\d+)\s*Mins?)?\s*(\d+)\s*Sec', str(duration_str), re.IGNORECASE)
    if not m:
        return None
    mins = int(m.group(1)) if m.group(1) else 0
    return mins * 60 + int(m.group(2))


def _find_previous_duration(wb, snum, today_sheet_name):
    """Finds this report's duration string in the most recent earlier day-sheet."""
    try:
        today_dt = datetime.strptime(today_sheet_name, "%d-%m-%Y")
    except ValueError:
        return None
    candidates = []
    for name in wb.sheetnames:
        if name in ("Performance", today_sheet_name):
            continue
        try:
            d = datetime.strptime(name, "%d-%m-%Y")
        except ValueError:
            continue
        if d < today_dt:
            candidates.append((d, name))
    if not candidates:
        return None
    candidates.sort()
    ws_prev = wb[candidates[-1][1]]
    for row in ws_prev.iter_rows(min_row=2, values_only=True):
        if row[0] == snum:
            return row[5]
    return None


def _format_comparison(today_str, prev_str):
    today_s = _parse_duration_seconds(today_str)
    prev_s = _parse_duration_seconds(prev_str)
    if today_s is None or not prev_s:
        return "—"
    pct = (today_s - prev_s) / prev_s * 100
    if pct > 0.05:
        return f"▲ {pct:.1f}% vs prev day"
    if pct < -0.05:
        return f"▼ {abs(pct):.1f}% vs prev day"
    return "→ 0% vs prev day"


def log_duration_to_excel(duration_str: str):
    """
    Opens Reports_Duration.xlsx. Appends a row to today's sheet (creating it
    from the Performance sheet's header if it doesn't exist yet). Optional —
    failure here never blocks the email send.
    """
    if DURATION_FILE is None:
        print("[Excel] DURATION_LOG_FILE not set — skipping duration log.")
        return
    wb = openpyxl.load_workbook(DURATION_FILE)
    source_sheet = wb["Performance"]

    headers = [cell.value for cell in source_sheet[1]] + ["vs Prev Day"]
    source_row = None
    for row in source_sheet.iter_rows(min_row=2, values_only=True):
        if row[0] == REPORT_SNUM:
            source_row = list(row)
            break

    if source_row is None:
        raise ValueError(f"S.No {REPORT_SNUM} not found in Performance sheet.")

    header_fill  = PatternFill("solid", start_color="1F4E79", end_color="1F4E79")
    header_font  = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    data_font    = Font(name="Calibri", size=11)
    center_align = Alignment(horizontal="center", vertical="center")
    left_align   = Alignment(horizontal="left",   vertical="center")
    thin         = Side(style="thin", color="BFBFBF")
    border       = Border(left=thin, right=thin, top=thin, bottom=thin)
    date_fill    = PatternFill("solid", start_color="E2EFDA", end_color="E2EFDA")
    dur_fill     = PatternFill("solid", start_color="FFF2CC", end_color="FFF2CC")
    col_widths   = [8, 15, 55, 18, 18, 20, 22]

    if SHEET_NAME in wb.sheetnames:
        ws = wb[SHEET_NAME]
        ws.cell(row=1, column=7, value="vs Prev Day").font = header_font
        ws.cell(row=1, column=7).fill = header_fill
        ws.cell(row=1, column=7).alignment = center_align
        ws.cell(row=1, column=7).border = border
        next_row = ws.max_row + 1
        for r in range(2, ws.max_row + 1):
            if ws.cell(row=r, column=1).value == REPORT_SNUM:
                ws.delete_rows(r, 1)
                next_row = ws.max_row + 1
                break
    else:
        ws = wb.create_sheet(title=SHEET_NAME)
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = border
        ws.row_dimensions[1].height = 20
        for i, w in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
        next_row = 2

    prev_duration = _find_previous_duration(wb, REPORT_SNUM, SHEET_NAME)
    comparison = _format_comparison(duration_str, prev_duration)

    data_row = source_row.copy()
    data_row[3] = MONTH_FROM
    data_row[4] = MONTH_TO
    data_row[5] = duration_str
    data_row.append(comparison)

    for col_idx, value in enumerate(data_row, start=1):
        cell = ws.cell(row=next_row, column=col_idx, value=value)
        cell.font = data_font
        cell.border = border
        if col_idx in (4, 5):
            cell.number_format = "DD/MM/YYYY"
            cell.fill = date_fill
            cell.alignment = center_align
        elif col_idx in (6, 7):
            cell.fill = dur_fill
            cell.alignment = center_align
        else:
            cell.alignment = left_align
    ws.row_dimensions[next_row].height = 18

    wb.save(DURATION_FILE)
    print(f"[Excel] Duration logged → sheet '{SHEET_NAME}' row {next_row} in {DURATION_FILE} ({comparison})")


# ── Main automation ──────────────────────────────────────────

def run():
    print(f"[Start] {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"[Report] {REPORT_NAME}  |  {MONTH_FROM} → {MONTH_TO}")

    with sync_playwright() as p:
        browser = p.chromium.launch(
            channel="chrome",   # uses installed Chrome
            headless=HEADLESS,
            downloads_path=str(DOWNLOAD_DIR),
            args=["--start-maximized", "--no-sandbox", "--disable-dev-shm-usage"]
        )

        context = browser.new_context(
            accept_downloads=True,
            viewport=None if not HEADLESS else {"width": 1280, "height": 900}
        )
        page = context.new_page()

        # ── 1. Open ERP
        print("[Nav] Opening ERP URL...")
        page.goto(ERP_URL, wait_until="domcontentloaded", timeout=60000)

        # ── 2. Login
        print("[Nav] Logging in...")
        page.wait_for_selector("input[name='ide_username']", state="visible", timeout=120000)

        page.click("input[name='ide_username']")
        page.fill("input[name='ide_username']", USERNAME)
        page.click("input[name='ide_password']")
        page.fill("input[name='ide_password']", PASSWORD)
        page.get_by_role("button", name="Login").click()

        try:
            page.wait_for_selector("text=already logged in", state="visible", timeout=8000)
            print("[Nav] Existing session dialog detected — confirming Yes.")
            page.get_by_role("button", name="Yes").click()
        except PlaywrightTimeout:
            pass

        print("[Nav] Login successful. Waiting for home screen...")
        wait_ready(page)

        # ── 3. Open main menu (hamburger icon)
        print("[Nav] Opening main menu...")
        page.click("a[data-id='mainMenu']", timeout=120000)
        page.wait_for_timeout(500)

        # ── 4. Navigate to Reports
        print("[Nav] Clicking Reports menu...")
        page.click("text=Reports", timeout=20000)
        page.wait_for_timeout(500)

        # ── 5. Report Management
        print("[Nav] Clicking Report Management...")
        page.click("text=Report Management", timeout=20000)
        page.wait_for_timeout(500)

        # ── 6. Manage Reports
        print("[Nav] Clicking Manage Reports...")
        page.click("text=Manage Reports", timeout=20000)
        wait_ready(page)

        # ── 7. Select Sales Register Mattresses
        print(f"[Nav] Selecting report: {REPORT_NAME}...")
        desc_cells = page.locator(f"text={REPORT_NAME}")
        report_link = None
        for i in range(desc_cells.count()):
            cell = desc_cells.nth(i)
            if not cell.is_visible():
                continue
            row_el = cell.evaluate_handle("e => e.closest('tr')").as_element()
            if row_el is None:
                continue
            link_el = row_el.query_selector(".x-column-ramcolink")
            if link_el is not None and link_el.is_visible():
                report_link = link_el
                break
        if report_link is None:
            raise RuntimeError(f"Could not find a clickable link for report '{REPORT_NAME}'.")

        dialog_frame_box = {"frame": None}

        def _on_navigated(fr):
            if "dw_htm_rt_ari_design" in fr.url:
                dialog_frame_box["frame"] = fr

        page.on("framenavigated", _on_navigated)
        try:
            report_link.click()
            print("[Nav] Waiting for report Parameters dialog frame...")
            deadline = time.time() + 280
            while dialog_frame_box["frame"] is None and time.time() < deadline:
                page.wait_for_timeout(200)
        finally:
            page.remove_listener("framenavigated", _on_navigated)

        dialog_frame = dialog_frame_box["frame"]
        if dialog_frame is None:
            page.screenshot(path="error_parameters_dialog.png")
            print("[ERROR] Parameters dialog frame never attached — see error_parameters_dialog.png")
            print("[DEBUG] Current frames:", [fr.url for fr in page.frames])
            raise RuntimeError("Report Parameters dialog frame never attached.")

        try:
            wait_for_visible_text(dialog_frame, ".x-btn", "Ok", timeout=150000)
        except PlaywrightTimeout:
            page.screenshot(path="error_parameters_dialog.png")
            print("[ERROR] Parameters dialog never rendered — see error_parameters_dialog.png")
            raise
        time.sleep(0.5)

        # ── 8. Set date parameters (current month, start to end).
        print(f"[Nav] Setting dates: {MONTH_FROM.strftime(DATE_FMT)} → {MONTH_TO.strftime(DATE_FMT)}")
        date_from_field = find_visible_input_by_value(
            dialog_frame, r"^\d{2}-\d{2}-\d{4}$", placeholder="DD-MM-YYYY"
        )
        if date_from_field is None:
            page.screenshot(path="error_date_field.png")
            values = dialog_frame.evaluate("""
                Array.from(document.querySelectorAll('input')).filter(el => {
                    const r = el.getBoundingClientRect();
                    return r.width > 0 && r.height > 0;
                }).map(el => el.value)
            """)
            print("[DEBUG] Visible input values in dialog frame:", values)
            raise RuntimeError("Could not find the Date From field.")
        print("[Nav] Clicking Date From field...")
        date_from_field.click()
        page.keyboard.press("Escape")
        time.sleep(0.3)
        print(f"[Nav] Typing Date From: {MONTH_FROM.strftime(DATE_FMT)}")
        page.keyboard.type(MONTH_FROM.strftime(DATE_FMT))
        time.sleep(0.3)
        print("[Nav] Tabbing to Date To field (1 tab)...")
        page.keyboard.press("Tab")
        time.sleep(0.1)
        print(f"[Nav] Typing Date To: {MONTH_TO.strftime(DATE_FMT)}")
        page.keyboard.type(MONTH_TO.strftime(DATE_FMT))
        time.sleep(0.3)
        page.keyboard.press("Tab")
        time.sleep(0.3)
        print("[Nav] Date entry complete.")

        actual_values = dialog_frame.evaluate("""
            () => {
                const els = Array.from(document.querySelectorAll('input')).filter(el => {
                    const r = el.getBoundingClientRect();
                    return r.width > 0 && r.height > 0;
                });
                return els.map(el => el.value).filter(v => /\\d{2}-\\d{2}-\\d{4}/.test(v));
            }
        """)
        print(f"[Nav] Date fields with date-shaped values now: {actual_values}")

        time.sleep(SETTLE_SECS)

        # ── 9. Click OK
        print("[Nav] Clicking OK...")
        wait_for_visible_text(dialog_frame, ".x-btn", "Ok", timeout=30000).click()
        wait_ready(page)

        # ── 9. Click Export — START TIMER
        print("[Export] Clicking Export... timer started.")
        start_time = time.time()
        page.keyboard.press("Escape")
        time.sleep(2)
        wait_ready(page)
        report_frame = None
        for fr in page.frames:
            try:
                el = fr.frame_element()
                if el and "ifReportDesigner" in (el.get_attribute("id") or ""):
                    report_frame = fr
                    break
            except Exception:
                continue
        if report_frame is None:
            raise RuntimeError("Could not find the report designer iframe.")

        report_frame.locator("#idExcel-btnIconEl").click()
        print("[Export] Clicked the Excel export icon.")

        time.sleep(5)
        page.wait_for_load_state("networkidle", timeout=30000)

        # ── 10. Confirm export format = Excel, click OK.
        print("[Export] Confirming export format dialog...")
        deadline = time.time() + 90
        ok_frame = None
        while time.time() < deadline and ok_frame is None:
            for fr in page.frames:
                if "exportdata_settings" in fr.url:
                    ok_frame = fr
                    break
            if ok_frame is None:
                page.wait_for_timeout(1000)
        if ok_frame is None:
            page.screenshot(path="debug_export_dialog.png", timeout=10000)
            print("[DEBUG] Frames at Ok-search timeout:", [fr.url for fr in page.frames])
            raise RuntimeError("Could not find the export settings dialog frame.")
        print(f"[DEBUG] Found export settings frame: {ok_frame.url}")
        wait_for_visible_text(ok_frame, ".x-btn", "Ok", timeout=30000)
        time.sleep(SETTLE_SECS)

        # ── 11. Click OK and capture the download directly.
        DOWNLOAD_TIMEOUT = 1800
        with page.expect_download(timeout=DOWNLOAD_TIMEOUT * 1000) as download_info:
            wait_for_visible_text(ok_frame, ".x-btn", "Ok", timeout=10000).click()
            print("[Export] Final OK clicked. Waiting for download...")
        download = download_info.value
        end_time = time.time()

        save_path = DOWNLOAD_DIR / f"{REPORT_NAME.replace(' ', '_')}_{TODAY.strftime('%Y%m%d_%H%M%S')}.xlsx"
        download.save_as(save_path)
        fix_xlsx_relationship_paths(save_path)
        new_file = save_path

        duration_sec = end_time - start_time
        duration_str = format_duration(duration_sec)

        print(f"[Done] File: {new_file.name}")
        print(f"[Done] Duration: {duration_str} ({duration_sec:.1f}s)")

        browser.close()

    # ── 12. Log to Excel (optional, non-fatal)
    try:
        log_duration_to_excel(duration_str)
    except Exception as exc:
        print(f"[Excel] Duration logging skipped (non-fatal): {exc}")

    print("[Complete] Download complete. Next step (email) handled by email_sender.py.")
    return new_file


if __name__ == "__main__":
    run()
